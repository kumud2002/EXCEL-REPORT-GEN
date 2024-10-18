from flask import Flask, render_template, request, send_file
import pandas as pd
from io import BytesIO
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
import smtplib
from email.message import EmailMessage
import ssl



app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Needed for session handling


def send_email_with_attachment(
        sender_email,
        sender_password,
        recipient_emails,
        subject,
        body,
        attachment_path,
        smtp_server,
        smtp_port,
        new_filename=None,
        use_ssl=True
):
    """
    Sends an email with an Excel attachment to multiple recipients.

    Parameters:
        sender_email (str): Sender's email address.
        sender_password (str): Sender's email password or app-specific password.
        recipient_emails (list): List of recipient email addresses.
        subject (str): Subject of the email.
        body (str): Body text of the email.
        attachment_path (str): File path to the Excel attachment.
        smtp_server (str): SMTP server address.
        smtp_port (int): SMTP server port.
        use_ssl (bool): Whether to use SSL for the SMTP connection.

    Returns:
        None
    """
    server = None  # Initialize server to None
    try:
        # Create EmailMessage object
        msg = EmailMessage()
        msg['From'] = sender_email
        msg['To'] = ', '.join(recipient_emails)
        msg['Subject'] = subject
        msg.set_content(body)

        # Read and add the Excel attachment
        with open(attachment_path, 'rb') as f:
            file_data = f.read()
            file_name = new_filename if new_filename else f.name.split('/')[-1]  # Use new filename if provided

        msg.add_attachment(
            file_data,
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=file_name
        )

        # Set up the SMTP server connection
        if use_ssl:
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, context=context)
            server.login(sender_email, sender_password)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls(context=ssl.create_default_context())
            server.login(sender_email, sender_password)

        # Send the email
        server.send_message(msg)

        print("Email sent successfully to:", recipient_emails)

    except Exception as e:
        print(f"Failed to send email: {e}")

    finally:
        if server:
            server.quit()  # Ensure server.quit() is called if server is defined


# Usage example

pathfor = None

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file1 = request.files['file1']
        selected_date = pd.to_datetime(request.form['date'])  # Get date input from the user


        # Process the uploaded file
        df = pd.read_excel(file1)

        # First DataFrame processing (from code1)
        # First DataFrame processing (from code1)
        df1 = df.copy()
        df1['Login Date'] = pd.to_datetime(df1['Login Date'], format='%d-%m-%Y')
        df1['Decision date'] = pd.to_datetime(df1['Decision date'], format='%d-%m-%Y')
        filtered_df = df1[df1['Login Date'] == selected_date]

        output_df1 = pd.DataFrame(columns=[
            'CCM', 'No (Login)', 'Val (In Lacs) (Login)',
            'No (Sanction)', 'Val (In Lacs) (Sanction)', 'No (Reject/Withdraw)',
            'Val (In Lacs) (Reject/Withdraw)', 'No (Decision)', 'Val (In Lacs) (Decision)',
            'No (Disbursed)', 'Val (In Lacs) (Disbursed)'
        ])

        groupeda = df1.groupby(['CCM'])
        grouped = filtered_df.groupby(['CCM'])

        for (ccm), group in groupeda:
            login_count = group[group['Login Date'] == selected_date]['Lead ID (Synofin)'].nunique()
            login_val = group[group['Login Date'] == selected_date]['Request Amount'].sum() / 100000
            sanction_count = group[group['Decision date'] == selected_date][group['Initial File Status (Credit)'] == 'Sanction'].shape[0]
            sanction_val = group[group['Decision date'] == selected_date][group['Initial File Status (Credit)'] == 'Sanction']['Sanction Amount'].sum() / 100000
            reject_count = group[group['Decision date'] == selected_date][group['Initial File Status (Credit)'] == 'Reject'].shape[0]
            reject_val = group[group['Decision date'] == selected_date][group['Initial File Status (Credit)'] == 'Reject']['Request Amount'].sum() / 100000
            decision_count = sanction_count + reject_count
            decision_val = sanction_val + reject_val
            disbursed_count = group[group['Decision date'] == selected_date][group['Disb. Date'].notnull()].shape[0]
            disbursed_val = group[group['Decision date'] == selected_date][group['Disb. Date'].notnull()]['Sanction Amount'].sum() / 100000

            output_df1 = output_df1.append({
                'CCM': ccm,
                'No (Login)': login_count,
                'Val (In Lacs) (Login)': login_val,
                'No (Sanction)': sanction_count,
                'Val (In Lacs) (Sanction)': sanction_val,
                'No (Reject/Withdraw)': reject_count,
                'Val (In Lacs) (Reject/Withdraw)': reject_val,
                'No (Decision)': decision_count,
                'Val (In Lacs) (Decision)': decision_val,
                'No (Disbursed)': disbursed_count,
                'Val (In Lacs) (Disbursed)': disbursed_val
            }, ignore_index=True)

        totals1 = {
            'CCM': 'Total',
            'No (Login)': output_df1['No (Login)'].sum(),
            'Val (In Lacs) (Login)': output_df1['Val (In Lacs) (Login)'].sum(),
            'No (Sanction)': output_df1['No (Sanction)'].sum(),
            'Val (In Lacs) (Sanction)': output_df1['Val (In Lacs) (Sanction)'].sum(),
            'No (Reject/Withdraw)': output_df1['No (Reject/Withdraw)'].sum(),
            'Val (In Lacs) (Reject/Withdraw)': output_df1['Val (In Lacs) (Reject/Withdraw)'].sum(),
            'No (Decision)': output_df1['No (Decision)'].sum(),
            'Val (In Lacs) (Decision)': output_df1['Val (In Lacs) (Decision)'].sum(),
            'No (Disbursed)': output_df1['No (Disbursed)'].sum(),
            'Val (In Lacs) (Disbursed)': output_df1['Val (In Lacs) (Disbursed)'].sum()
        }

        output_df1 = output_df1.append(totals1, ignore_index=True)

        # Second DataFrame processing (from code2)
        output_df2 = pd.DataFrame(columns=[
            'CCM', 'No (Spill File)', 'Val (In Lacs) (Spill File)', 'No (Fresh Login)', 'Val (In Lacs) (Fresh Login)',
            'No (Total File)', 'Val (In Lacs) (Total File)', 'No (Sanction/Disbursed)', 'Val (In Lacs) (Sanction/Disbursed)',
            'No (Reject/Withdraw)', 'Val (In Lacs) (Reject/Withdraw)', 'No (Recommend)', 'Val (In Lacs) (Recommend)',
            'No (Query-Sales)', 'Val (In Lacs) (Query-Sales)', 'No (WIP-Credit)', 'Val (In Lacs) (WIP-Credit)',
            'No (Visit Pending)', 'Val (In Lacs) (Visit Pending)'
        ])

        user_date = selected_date
        current_month = user_date.strftime('%b')

        # Collect all previous months
        all_previous_months = [user_date.replace(day=1) - pd.DateOffset(months=i) for i in range(1, user_date.month)]
        previous_months_str = [month.strftime('%b') for month in all_previous_months]

        for ccm, group in df.groupby(['CCM']):
            aug_group = group[group['MONTH'] == current_month]
            naug_group = group[group['MONTH'].isin(previous_months_str)]

            spfno = len(naug_group)
            spfval = naug_group['Request Amount'].sum() / 100000
            frshlg = len(aug_group)
            frshval = aug_group['Request Amount'].sum() / 100000
            totalno = spfno + frshlg
            totalval = spfval + frshval
            sandis_count = group[group['Initial File Status (Credit)'] == 'Sanction'].shape[0] + group[group['Initial File Status (Credit)'] == 'Disbursed'].shape[0]
            sandis_val = group[group['Initial File Status (Credit)'] == 'Sanction']['Sanction Amount'].sum() / 100000 + group[group['Initial File Status (Credit)'] == 'Disbursed']['Sanction Amount'].sum() / 100000
            reject_count = group[group['Initial File Status (Credit)'] == 'Reject'].shape[0]
            reject_val = group[group['Initial File Status (Credit)'] == 'Reject']['Request Amount'].sum() / 100000
            rec_count = group[group['Initial File Status (Credit)'] == 'Recommend'].shape[0]
            rec_val = group[group['Initial File Status (Credit)'] == 'Recommend']['Request Amount'].sum() / 100000
            qs_count = group[group['Initial File Status (Credit)'] == 'Query- Sales'].shape[0]
            qs_val = group[group['Initial File Status (Credit)'] == 'Query- Sales']['Request Amount'].sum() / 100000
            wip_count = group[group['Initial File Status (Credit)'] == 'WIP- Credit'].shape[0]
            wip_val = group[group['Initial File Status (Credit)'] == 'WIP- Credit']['Request Amount'].sum() / 100000
            vp_count = group[group['Initial File Status (Credit)'] == 'Visit Pending'].shape[0]
            vp_val = group[group['Initial File Status (Credit)'] == 'Visit Pending']['Request Amount'].sum() / 100000

            output_df2 = output_df2.append({
                'CCM': ccm,
                'No (Spill File)': spfno,
                'Val (In Lacs) (Spill File)': spfval,
                'No (Fresh Login)': frshlg,
                'Val (In Lacs) (Fresh Login)': frshval,
                'No (Total File)': totalno,
                'Val (In Lacs) (Total File)': totalval,
                'No (Sanction/Disbursed)': sandis_count,
                'Val (In Lacs) (Sanction/Disbursed)': sandis_val,
                'No (Reject/Withdraw)': reject_count,
                'Val (In Lacs) (Reject/Withdraw)': reject_val,
                'No (Recommend)': rec_count,
                'Val (In Lacs) (Recommend)': rec_val,
                'No (Query-Sales)': qs_count,
                'Val (In Lacs) (Query-Sales)': qs_val,
                'No (WIP-Credit)': wip_count,
                'Val (In Lacs) (WIP-Credit)': wip_val,
                'No (Visit Pending)': vp_count,
                'Val (In Lacs) (Visit Pending)': vp_val
            }, ignore_index=True)

    # Assuming output_df2 is already created and populated

    # Calculate totals for each numeric column
        totals2 = {
        'CCM': 'Total',
        'No (Spill File)': output_df2['No (Spill File)'].sum(),
        'Val (In Lacs) (Spill File)': output_df2['Val (In Lacs) (Spill File)'].sum(),
        'No (Fresh Login)': output_df2['No (Fresh Login)'].sum(),
        'Val (In Lacs) (Fresh Login)': output_df2['Val (In Lacs) (Fresh Login)'].sum(),
        'No (Total File)': output_df2['No (Total File)'].sum(),
        'Val (In Lacs) (Total File)': output_df2['Val (In Lacs) (Total File)'].sum(),
        'No (Sanction/Disbursed)': output_df2['No (Sanction/Disbursed)'].sum(),
        'Val (In Lacs) (Sanction/Disbursed)': output_df2['Val (In Lacs) (Sanction/Disbursed)'].sum(),
        'No (Reject/Withdraw)': output_df2['No (Reject/Withdraw)'].sum(),
        'Val (In Lacs) (Reject/Withdraw)': output_df2['Val (In Lacs) (Reject/Withdraw)'].sum(),
        'No (Recommend)': output_df2['No (Recommend)'].sum(),
        'Val (In Lacs) (Recommend)': output_df2['Val (In Lacs) (Recommend)'].sum(),
        'No (Query-Sales)': output_df2['No (Query-Sales)'].sum(),
        'Val (In Lacs) (Query-Sales)': output_df2['Val (In Lacs) (Query-Sales)'].sum(),
        'No (WIP-Credit)': output_df2['No (WIP-Credit)'].sum(),
        'Val (In Lacs) (WIP-Credit)': output_df2['Val (In Lacs) (WIP-Credit)'].sum(),
        'No (Visit Pending)': output_df2['No (Visit Pending)'].sum(),
        'Val (In Lacs) (Visit Pending)': output_df2['Val (In Lacs) (Visit Pending)'].sum()
    }

    # Append the totals row to output_df2
        output_df2 = output_df2.append(totals2, ignore_index=True)

        output_df3 = pd.DataFrame(columns=[
            'CBM', 'No (Login)', 'Val (In Lacs) (Login)',
            'No (Sanction)', 'Val (In Lacs) (Sanction)', 'No (Reject/Withdraw)',
            'Val (In Lacs) (Reject/Withdraw)', 'No (Decision)', 'Val (In Lacs) (Decision)',
            'No (Disbursed)', 'Val (In Lacs) (Disbursed)'
        ])

        grouped1 = df1.groupby(['CBM'])

        for (cbm), group in grouped1:
            login_count = group[group['Login Date']== selected_date]['Lead ID (Synofin)'].nunique()
            login_val = group[group['Login Date']== selected_date]['Request Amount'].sum() / 100000
            sanction_count = group[group['Decision date']== selected_date][group['Initial File Status (Credit)'] == 'Sanction'].shape[0]
            sanction_val = group[group['Decision date']== selected_date][group['Initial File Status (Credit)'] == 'Sanction']['Sanction Amount'].sum() / 100000
            reject_count = group[group['Decision date']== selected_date][group['Initial File Status (Credit)'] == 'Reject'].shape[0]
            reject_val = group[group['Decision date']== selected_date][group['Initial File Status (Credit)'] == 'Reject']['Request Amount'].sum() / 100000
            decision_count = sanction_count + reject_count
            decision_val = sanction_val + reject_val
            disbursed_count = group[group['Decision date']== selected_date][group['Disb. Date'].notnull()].shape[0]
            disbursed_val = group[group['Decision date']== selected_date][group['Disb. Date'].notnull()]['Sanction Amount'].sum() / 100000

            output_df3 = output_df3.append({
                'CBM': cbm,
                'No (Login)': login_count,
                'Val (In Lacs) (Login)': login_val,
                'No (Sanction)': sanction_count,
                'Val (In Lacs) (Sanction)': sanction_val,
                'No (Reject/Withdraw)': reject_count,
                'Val (In Lacs) (Reject/Withdraw)': reject_val,
                'No (Decision)': decision_count,
                'Val (In Lacs) (Decision)': decision_val,
                'No (Disbursed)': disbursed_count,
                'Val (In Lacs) (Disbursed)': disbursed_val
            }, ignore_index=True)

        totals3 = {
            'CBM': 'Total',
            'No (Login)': output_df3['No (Login)'].sum(),
            'Val (In Lacs) (Login)': output_df3['Val (In Lacs) (Login)'].sum(),
            'No (Sanction)': output_df3['No (Sanction)'].sum(),
            'Val (In Lacs) (Sanction)': output_df3['Val (In Lacs) (Sanction)'].sum(),
            'No (Reject/Withdraw)': output_df3['No (Reject/Withdraw)'].sum(),
            'Val (In Lacs) (Reject/Withdraw)': output_df3['Val (In Lacs) (Reject/Withdraw)'].sum(),
            'No (Decision)': output_df3['No (Decision)'].sum(),
            'Val (In Lacs) (Decision)': output_df3['Val (In Lacs) (Decision)'].sum(),
            'No (Disbursed)': output_df3['No (Disbursed)'].sum(),
            'Val (In Lacs) (Disbursed)': output_df3['Val (In Lacs) (Disbursed)'].sum()
        }
        # Append the totals row to output_df2
        output_df3 = output_df3.append(totals3, ignore_index=True)

        # Second DataFrame processing (from code2)
        output_df4 = pd.DataFrame(columns=[
            'CBM', 'No (Spill File)', 'Val (In Lacs) (Spill File)', 'No (Fresh Login)', 'Val (In Lacs) (Fresh Login)',
            'No (Total File)', 'Val (In Lacs) (Total File)', 'No (Sanction/Disbursed)',
            'Val (In Lacs) (Sanction/Disbursed)',
            'No (Reject/Withdraw)', 'Val (In Lacs) (Reject/Withdraw)', 'No (Recommend)', 'Val (In Lacs) (Recommend)',
            'No (Query-Sales)', 'Val (In Lacs) (Query-Sales)', 'No (WIP-Credit)', 'Val (In Lacs) (WIP-Credit)',
            'No (Visit Pending)', 'Val (In Lacs) (Visit Pending)'
        ])


        for cbm, group in df.groupby(['CBM']):
            aug_group = group[group['MONTH'] == current_month]
            naug_group = group[group['MONTH'].isin(previous_months_str)]

            spfno = len(naug_group)
            spfval = naug_group['Request Amount'].sum() / 100000
            frshlg = len(aug_group)
            frshval = aug_group['Request Amount'].sum() / 100000
            totalno = spfno + frshlg
            totalval = spfval + frshval
            sandis_count = group[group['Initial File Status (Credit)'] == 'Sanction'].shape[0] + group[group['Initial File Status (Credit)'] == 'Disbursed'].shape[0]
            sandis_val = group[group['Initial File Status (Credit)'] == 'Sanction']['Sanction Amount'].sum() / 100000 + group[group['Initial File Status (Credit)'] == 'Disbursed']['Sanction Amount'].sum() / 100000
            reject_count = group[group['Initial File Status (Credit)'] == 'Reject'].shape[0]
            reject_val = group[group['Initial File Status (Credit)'] == 'Reject']['Request Amount'].sum() / 100000
            rec_count = group[group['Initial File Status (Credit)'] == 'Recommend'].shape[0]
            rec_val = group[group['Initial File Status (Credit)'] == 'Recommend']['Request Amount'].sum() / 100000
            qs_count = group[group['Initial File Status (Credit)'] == 'Query- Sales'].shape[0]
            qs_val = group[group['Initial File Status (Credit)'] == 'Query- Sales']['Request Amount'].sum() / 100000
            wip_count = group[group['Initial File Status (Credit)'] == 'WIP- Credit'].shape[0]
            wip_val = group[group['Initial File Status (Credit)'] == 'WIP- Credit']['Request Amount'].sum() / 100000
            vp_count = group[group['Initial File Status (Credit)'] == 'Visit Pending'].shape[0]
            vp_val = group[group['Initial File Status (Credit)'] == 'Visit Pending']['Request Amount'].sum() / 100000

            output_df4 = output_df4.append({
                'CBM': cbm,
                'No (Spill File)': spfno,
                'Val (In Lacs) (Spill File)': spfval,
                'No (Fresh Login)': frshlg,
                'Val (In Lacs) (Fresh Login)': frshval,
                'No (Total File)': totalno,
                'Val (In Lacs) (Total File)': totalval,
                'No (Sanction/Disbursed)': sandis_count,
                'Val (In Lacs) (Sanction/Disbursed)': sandis_val,
                'No (Reject/Withdraw)': reject_count,
                'Val (In Lacs) (Reject/Withdraw)': reject_val,
                'No (Recommend)': rec_count,
                'Val (In Lacs) (Recommend)': rec_val,
                'No (Query-Sales)': qs_count,
                'Val (In Lacs) (Query-Sales)': qs_val,
                'No (WIP-Credit)': wip_count,
                'Val (In Lacs) (WIP-Credit)': wip_val,
                'No (Visit Pending)': vp_count,
                'Val (In Lacs) (Visit Pending)': vp_val
            }, ignore_index=True)

        # Assuming output_df2 is already created and populated

        # Calculate totals for each numeric column
        totals4 = {
            'CBM': 'Total',
            'No (Spill File)': output_df4['No (Spill File)'].sum(),
            'Val (In Lacs) (Spill File)': output_df4['Val (In Lacs) (Spill File)'].sum(),
            'No (Fresh Login)': output_df4['No (Fresh Login)'].sum(),
            'Val (In Lacs) (Fresh Login)': output_df4['Val (In Lacs) (Fresh Login)'].sum(),
            'No (Total File)': output_df4['No (Total File)'].sum(),
            'Val (In Lacs) (Total File)': output_df4['Val (In Lacs) (Total File)'].sum(),
            'No (Sanction/Disbursed)': output_df4['No (Sanction/Disbursed)'].sum(),
            'Val (In Lacs) (Sanction/Disbursed)': output_df4['Val (In Lacs) (Sanction/Disbursed)'].sum(),
            'No (Reject/Withdraw)': output_df4['No (Reject/Withdraw)'].sum(),
            'Val (In Lacs) (Reject/Withdraw)': output_df4['Val (In Lacs) (Reject/Withdraw)'].sum(),
            'No (Recommend)': output_df4['No (Recommend)'].sum(),
            'Val (In Lacs) (Recommend)': output_df4['Val (In Lacs) (Recommend)'].sum(),
            'No (Query-Sales)': output_df4['No (Query-Sales)'].sum(),
            'Val (In Lacs) (Query-Sales)': output_df4['Val (In Lacs) (Query-Sales)'].sum(),
            'No (WIP-Credit)': output_df4['No (WIP-Credit)'].sum(),
            'Val (In Lacs) (WIP-Credit)': output_df4['Val (In Lacs) (WIP-Credit)'].sum(),
            'No (Visit Pending)': output_df4['No (Visit Pending)'].sum(),
            'Val (In Lacs) (Visit Pending)': output_df4['Val (In Lacs) (Visit Pending)'].sum()
        }
        # Append the totals row to output_df2
        output_df4 = output_df4.append(totals4, ignore_index=True)

        output_df5 = pd.DataFrame(columns=[
            'CCM', 'Visit Pending', 'WIP- Credit', 'Recommend', 'Reject', 'Query- Sales',
            'Sanction', 'Disbursed', 'Grand Total'
        ])


        for (ccm), group in df.groupby('CCM'):
            vp_count5 = group[group['Initial File Status (Credit)'] == 'Visit Pending'].shape[0]
            wip_count5 = group[group['Initial File Status (Credit)'] == 'WIP- Credit'].shape[0]
            recc_count5 = group[group['Initial File Status (Credit)'] == 'Recommend'].shape[0]
            reject_count5 = group[group['Initial File Status (Credit)'] == 'Reject'].shape[0]
            qs_count5 = group[group['Initial File Status (Credit)'] == 'Query- Sales'].shape[0]
            sanction_count5 = group[group['Initial File Status (Credit)'] == 'Sanction'].shape[0]
            disb_count5 = group[group['Initial File Status (Credit)'] == 'Disbursed'].shape[0]
            gt_count= vp_count5+wip_count5+recc_count5+reject_count5+sanction_count5+qs_count5+disb_count5

            output_df5 = output_df5.append({
                'CCM': ccm,
                'Visit Pending': vp_count5,
                'WIP- Credit': wip_count5,
                'Recommend': recc_count5,
                'Reject': reject_count5,
                'Query- Sales':qs_count5,
                'Sanction': sanction_count5,
                'Disbursed': disb_count5,
                'Grand Total': gt_count,

            }, ignore_index=True)

        totals5 = {
            'CCM': 'Grand Total',
            'Visit Pending': output_df5['Visit Pending'].sum(),
            'WIP- Credit': output_df5['WIP- Credit'].sum(),
            'Recommend': output_df5['Recommend'].sum(),
            'Reject': output_df5['Reject'].sum(),
            'Query- Sales': output_df5['Query- Sales'].sum(),
            'Sanction': output_df5['Sanction'].sum(),
            'Disbursed': output_df5['Disbursed'].sum(),
            'Grand Total': output_df5['Grand Total'].sum(),

        }

        # Append the totals row to output_df2
        output_df5 = output_df5.append(totals5, ignore_index=True)

        df_selected = df[['CCM', 'Login to PD TAT']]
        df_selected['Login to PD TAT'] = pd.to_numeric(df_selected['Login to PD TAT'], errors='coerce')
        avg_tat_by_ccm = df_selected.groupby('CCM')['Login to PD TAT'].mean().reset_index()

        totals6 = {
            'CCM': 'Grand Total',
            'Login to PD TAT': avg_tat_by_ccm['Login to PD TAT'].mean()

        }
        # Append the totals row to output_df2
        avg_tat_by_ccm = avg_tat_by_ccm.append(totals6, ignore_index=True)

        df_selected = df[["Visit Official Name\n(Credit Part)", 'Login to PD TAT']]
        df_selected['Login to PD TAT'] = pd.to_numeric(df_selected['Login to PD TAT'], errors='coerce')
        avg_tat_by_bcm = df_selected.groupby("Visit Official Name\n(Credit Part)")['Login to PD TAT'].mean().reset_index()

        totals7 = {
            "Visit Official Name\n(Credit Part)": 'Grand Total',
            'Login to PD TAT': avg_tat_by_bcm['Login to PD TAT'].mean()

        }
        # Append the totals row to output_df2
        avg_tat_by_bcm = avg_tat_by_bcm.append(totals7, ignore_index=True)


        # Save dataframes to a temporary Excel file with multiple sheets
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        with pd.ExcelWriter(temp_file.name, engine='xlsxwriter') as writer:
            output_df1.to_excel(writer, index=False, sheet_name='CCM FTD')
            output_df2.to_excel(writer, index=False, sheet_name='CCM MTD')
            output_df3.to_excel(writer, index=False, sheet_name='CBM FTD')
            output_df4.to_excel(writer, index=False, sheet_name='CBM MTD')
            output_df5.to_excel(writer, index=False, sheet_name='SWS')
            avg_tat_by_ccm.round(3).to_excel(writer, index=False, sheet_name='CCM TAT Avg')
            avg_tat_by_bcm.round(3).to_excel(writer, index=False, sheet_name='BCM TAT Avg')

        temp_file_path = temp_file.name

        # Load the workbook and apply formatting
        workbook = load_workbook(temp_file_path)
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        header_fill = PatternFill(start_color="9C0201", end_color="9C0201", fill_type="solid")

        total_font = Font(bold=True, color="FFFFFF", size=12, name='Calibri')
        total_fill = PatternFill(start_color="3C3B6E", end_color="3C3B6E", fill_type="solid")

        cell_font = Font(size=11, name='Calibri')
        alignment = Alignment(horizontal='center', vertical='center')

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define off-white fill for non-total, non-header cells
        off_white_fill = PatternFill(start_color="F7F4EB", end_color="F7F4EB", fill_type="solid")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Apply styles to the header
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border  # Apply border to header cells

            # Apply styles to the cells
            for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
                for cell in row:
                    cell.font = cell_font
                    cell.alignment = alignment
                    cell.border = thin_border  # Apply border to all cells

                # Apply formatting to the "Total" row
                if row[0].value == 'Total' or row[0].value == 'Grand Total':
                    for cell in row:
                        cell.font = total_font
                        cell.fill = total_fill
                        cell.border = thin_border  # Apply border to total cells
                else:
                    for cell in row:
                        cell.fill = off_white_fill  # Apply off-white color to non-total, non-header cells
                        cell.border = thin_border  # Ensure border is applied to non-total cells

            # Adjust column widths based on the maximum length of the content
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (A, B, C, etc.)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add some padding
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the modified workbook back to the same file
        workbook.save(temp_file_path)
        global pathfor
        pathfor = temp_file_path

        # Return the modified Excel file as a download
        return send_file(temp_file_path, as_attachment=True, download_name= selected_date.strftime("%d-%m-%Y")+' '+'Updated MIS Report.xlsx')

    return render_template('index.html')

@app.route('/send_email', methods=['POST'])
def trigger_send():
    # selected_date = pd.to_datetime(request.form['date'])
    subject = 'Here is Daily Your MIS Excel Report'
    body = 'Please find the attached Excel report.'
    sender_email = 'mananya.gaur@paisabuddy.com'
    sender_password = 'Pb@101010'  # Be cautious with passwords
    recipient_emails = ['tanishka.narula@paisabuddy.com','kumud.jain@paisabuddy.com']

    send_email_with_attachment(sender_email, sender_password, recipient_emails, subject, body, pathfor, 'smtp.zoho.com', 465, 'Updated MIS Report.xlsx', use_ssl=True)

    return render_template('index.html')



if __name__ == '__main__':
    app.run(debug=True)
